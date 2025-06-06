import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split, StratifiedKFold, GridSearchCV, RandomizedSearchCV, learning_curve
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import MinMaxScaler
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import (accuracy_score, roc_auc_score, confusion_matrix, classification_report,
                             precision_score, recall_score, f1_score, roc_curve)
from scipy import stats
import seaborn as sns
from openpyxl import load_workbook
import xlsxwriter
import os
from sklearn.metrics import roc_curve, auc, precision_recall_curve, average_precision_score
from sklearn.calibration import calibration_curve
from imblearn.over_sampling import RandomOverSampler
from imblearn.combine import SMOTEENN
from imblearn.over_sampling import SMOTE
from collections import Counter
import shap
from sklearn.metrics import roc_curve, confusion_matrix, accuracy_score, recall_score, precision_score, f1_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import make_scorer, roc_auc_score, recall_score, confusion_matrix
import joblib
from typing import List, Optional
from sklearn.feature_selection import RFE
from mrmr import mrmr_classif

from collections import defaultdict
from sklearn.model_selection import KFold, StratifiedKFold
from sklearn.model_selection import cross_val_score

from scipy.ndimage import gaussian_filter1d
from sklearn.calibration import CalibratedClassifierCV
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split, StratifiedKFold, GridSearchCV, RandomizedSearchCV, learning_curve
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import MinMaxScaler
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import (accuracy_score, roc_auc_score, confusion_matrix, classification_report,
                             precision_score, recall_score, f1_score, roc_curve)
from scipy import stats
import seaborn as sns
from openpyxl import load_workbook
import xlsxwriter
import os
from sklearn.metrics import roc_curve, auc, precision_recall_curve, average_precision_score
from sklearn.calibration import calibration_curve
from imblearn.over_sampling import RandomOverSampler
from imblearn.combine import SMOTEENN
from imblearn.over_sampling import SMOTE
from collections import Counter
import shap
from sklearn.metrics import roc_curve, confusion_matrix, accuracy_score, recall_score, precision_score, f1_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import make_scorer, roc_auc_score, recall_score, confusion_matrix
import joblib
from typing import List, Optional
from sklearn.feature_selection import RFE
from sklearn.neural_network import MLPClassifier
from scipy.ndimage import gaussian_filter1d
from typing import List, Optional
from mrmr import mrmr_classif
from sklearn.model_selection import KFold
from collections import defaultdict
from sklearn.model_selection import KFold, StratifiedKFold
from sklearn.metrics import roc_curve, roc_auc_score
import joblib
import matplotlib.pyplot as plt
from dcurves import dca, plot_graphs, load_test_data
from src.visualization.plotting import *
from sklearn.metrics import recall_score

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split, StratifiedKFold, GridSearchCV, RandomizedSearchCV, learning_curve
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import MinMaxScaler
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import (accuracy_score, roc_auc_score, confusion_matrix, classification_report,
                             precision_score, recall_score, f1_score, roc_curve)
from scipy import stats
import seaborn as sns
from openpyxl import load_workbook
import xlsxwriter
import os
from sklearn.metrics import roc_curve, auc, precision_recall_curve, average_precision_score
from sklearn.calibration import calibration_curve
from imblearn.over_sampling import RandomOverSampler
from imblearn.combine import SMOTEENN
from imblearn.over_sampling import SMOTE
from collections import Counter
import shap
from sklearn.metrics import roc_curve, confusion_matrix, accuracy_score, recall_score, precision_score, f1_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import make_scorer, roc_auc_score, recall_score, confusion_matrix
import joblib
from typing import List, Optional
from sklearn.feature_selection import RFE
from sklearn.neural_network import MLPClassifier
from scipy.ndimage import gaussian_filter1d
from typing import List, Optional
from mrmr import mrmr_classif
from sklearn.model_selection import KFold
from collections import defaultdict
from sklearn.model_selection import KFold, StratifiedKFold
from sklearn.metrics import roc_curve, roc_auc_score
import joblib
import matplotlib.pyplot as plt
from dcurves import dca, plot_graphs, load_test_data
from src.visualization.plotting import *




SAVE_PLOTS = True


def get_classifiers_simple():
    """
    Returns a dictionary of classifiers with their hyperparameter grids to be evaluated.
    """
    return {
        'RandomForest': RandomForestClassifier(class_weight='balanced'),
        'SVM': SVC(probability=True, class_weight='balanced'),
        'LogisticRegression': LogisticRegression(class_weight='balanced'),
        'NaiveBayes': GaussianNB()
    }


def get_classifiers():
    """
    Returns a dictionary of classifiers with their hyperparameter grids to be evaluated.
    """
    return {
        'RandomForest': (RandomForestClassifier(class_weight='balanced', random_state=42), {
            'n_estimators': [25, 50],
            'max_features': ['sqrt'],
            'max_depth': [2, 3, 4],
            'min_samples_split': [10, 15, 20],
            'min_samples_leaf': [4, 6, 8],
            'criterion': ['gini'],
            'bootstrap': [True]
        }),
        'SVM': (SVC(probability=True, class_weight='balanced', kernel='rbf'), {
        }),
        # 'SVM': (SVC(probability=True, class_weight='balanced', kernel='rbf', random_state=42), {
        #     'C': [0.1, 1, 10, 100],  # Focus on relevant C values
        #     'gamma': [0.01, 0.001, 0.0001]  # Try a range of gamma values
        # }),

        # 'SVM': (Pipeline([
        #         ('scaler', StandardScaler()),
        #         ('svc', SVC(probability=True, class_weight='balanced', kernel='rbf', random_state=42))
        #     ]), {
        #     'svc__C': [0.01, 0.1, 1, 10, 100],
        #     'svc__gamma': [0.001, 0.01, 0.1, 1]
        # }),
        'LogisticRegression': (Pipeline([
                ('scaler', StandardScaler()),
                ('lr', LogisticRegression(class_weight='balanced', random_state=42, solver='liblinear'))
            ]), {
            'lr__penalty': ['l1', 'l2'],
            'lr__C': [0.01, 0.1, 1, 10]
        }),
        'NaiveBayes': (GaussianNB(var_smoothing=1e-9), {
        })
    }



def get_classifiers1():
    """
    Returns a dictionary of classifiers with their hyperparameter grids to be evaluated.
    """
    return {
        'RandomForest': (RandomForestClassifier(class_weight='balanced', random_state=42), {
            'n_estimators': [25, 50],
            'max_features': ['sqrt'],
            'max_depth': [2, 3, 4],
            'min_samples_split': [10, 15, 20],
            'min_samples_leaf': [4, 6, 8],
            'criterion': ['gini'],
            'bootstrap': [True]
        })
    }

# def compute_metrics(y_true, y_pred, y_pred_prob, target_sensitivity=None):
#     """
#     Compute evaluation metrics and their confidence intervals, with the option
#     to fix sensitivity.
#
#     Parameters:
#     y_true (array-like): True labels.
#     y_pred (array-like): Predicted labels.
#     y_pred_prob (array-like): Predicted probabilities.
#     target_sensitivity (float, optional): Target sensitivity value. If None,
#                                          the optimal threshold based on Youden's
#                                          Index is used.
#
#     Returns:
#     dict: Evaluation metrics and their confidence intervals.
#     """
#
#     # Compute Youden's Index to find the optimal threshold
#     fpr, tpr, thresholds = roc_curve(y_true, y_pred_prob)
#     youdens_index = tpr - fpr
#     optimal_idx = np.argmax(youdens_index)
#     optimal_threshold = thresholds[optimal_idx]
#
#     print(f"optimal_threshold is {optimal_threshold}")
#
#
#     # If target sensitivity is specified, find the threshold that achieves it
#     if target_sensitivity is not None:
#         sensitivity_idx = np.argmin(np.abs(tpr - target_sensitivity))
#         threshold = thresholds[sensitivity_idx]
#         print(f"target_sensitivity is {target_sensitivity} and threshold is {threshold}")
#
#
#     # Use the appropriate threshold for new predictions
#     if target_sensitivity is not None:
#         y_pred_optimal = (y_pred_prob >= threshold).astype(int)
#     else:
#         y_pred_optimal = (y_pred_prob >= optimal_threshold).astype(int)
#
#     print(f"y_pred_optimal is {y_pred_optimal}")
#
#     # Compute metrics using optimal threshold predictions
#     accuracy = accuracy_score(y_true, y_pred_optimal)
#     roc_auc = roc_auc_score(y_true, y_pred_prob) if y_pred_prob is not None else None
#     cm = confusion_matrix(y_true, y_pred_optimal)
#     tn, fp, fn, tp = cm.ravel()
#
#     print(f"tn is {tn}")
#     print(f"fp is {fp}")
#     print(f"fn is {fn}")
#     print(f"tp is {tp}")
#
#
#     specificity = tn / (tn + fp) if (tn + fp) else 0
#     sensitivity = tp / (tp + fn) if (tp + fn) else 0
#     ppv = precision_score(y_true, y_pred_optimal)
#     npv = tn / (tn + fn) if (tn + fn) else 0
#     f1 = f1_score(y_true, y_pred_optimal)
#
#     print(f"sensitivity is {sensitivity}")
#
#     metrics = {
#         'accuracy': accuracy,
#         'roc_auc': roc_auc,
#         'specificity': specificity,
#         'sensitivity': sensitivity,
#         'ppv': ppv,
#         'npv': npv,
#         'f1_score': f1
#     }
#
#     ci = {}
#     for metric, value in metrics.items():
#         if value is not None:
#             ci[metric] = compute_confidence_interval(value, y_true.size)
#
#     return metrics, ci



# def compute_metrics(y_true, y_pred_prob, predefined_thresh = None, target_sensitivity=0.8):
#     """
#     Compute evaluation metrics and their confidence intervals,
#     with precise sensitivity targeting.
#
#     Parameters:
#     y_true (array-like): True labels.
#     y_pred (array-like): Predicted labels.
#     y_pred_prob (array-like): Predicted probabilities.
#     target_sensitivity (float, optional): Target sensitivity value.
#
#     Returns:
#     dict: Evaluation metrics and their confidence intervals.
#     """
#     from sklearn.metrics import roc_curve, confusion_matrix, precision_score, accuracy_score, roc_auc_score, f1_score
#
#     # Find the threshold that most precisely matches target sensitivity
#     def find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity):
#         fpr, tpr, thresholds = roc_curve(y_true, y_pred_prob)
#
#         ##-----------------------------------
#         ## Find the threshold that gives sensitivity closest to target
#         best_threshold = None
#         min_diff = float('inf')
#         for i, (threshold, current_tpr) in enumerate(zip(thresholds, tpr)):
#             y_pred_at_threshold = (y_pred_prob >= threshold).astype(int)
#             current_sensitivity = recall_score(y_true, y_pred_at_threshold)
#
#             diff = abs(current_sensitivity - target_sensitivity)
#             if diff < min_diff:
#                 min_diff = diff
#                 best_threshold = threshold
#
#         ## -----------------------------------
#         ## Find the threshold that gives a higher than target
#         # best_threshold = None
#         # for i, (threshold, current_tpr) in enumerate(zip(thresholds, tpr)):
#         #     y_pred_at_threshold = (y_pred_prob >= threshold).astype(int)
#         #     current_sensitivity = recall_score(y_true, y_pred_at_threshold)
#         #
#         #     if current_sensitivity >= target_sensitivity:
#         #         best_threshold = threshold
#         #         break
#
#
#         # -----------------------------------
#         return best_threshold
#
#
#
#     threshold = 0.5
#
#     # Find the optimal threshold
#     if target_sensitivity:
#         threshold = find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity)
#
#     if predefined_thresh:
#         threshold = predefined_thresh
#
#     y_pred_optimal = (y_pred_prob >= threshold).astype(int)
#
#     # Compute metrics using optimal threshold predictions
#     cm = confusion_matrix(y_true, y_pred_optimal)
#     tn, fp, fn, tp = cm.ravel()
#
#
#     metrics = {
#         'accuracy': accuracy_score(y_true, y_pred_optimal),
#         'roc_auc': roc_auc_score(y_true, y_pred_prob) if y_pred_prob is not None else None,
#         'specificity': tn / (tn + fp) if (tn + fp) else 0,
#         'sensitivity': tp / (tp + fn) if (tp + fn) else 0,
#         'ppv': precision_score(y_true, y_pred_optimal),
#         'npv': tn / (tn + fn) if (tn + fn) else 0,
#         'f1_score': f1_score(y_true, y_pred_optimal),
#         'tn': tn,
#         'fp': fp,
#         'fn': fn,
#         'tp': tp,
#         'threshold': threshold
#     }
#
#     ci = {}
#     for metric, value in metrics.items():
#         if value is not None:
#             ci[metric] = compute_confidence_interval(value, y_true.size)
#
#     return metrics, ci, y_pred_optimal

def compute_metrics(y_true, y_pred_prob, predefined_thresh = None, target_sensitivity=0.8):
    """
    Compute evaluation metrics and their confidence intervals,
    with precise sensitivity targeting.

    Parameters:
    y_true (array-like): True labels.
    y_pred (array-like): Predicted labels.
    y_pred_prob (array-like): Predicted probabilities.
    target_sensitivity (float, optional): Target sensitivity value.

    Returns:
    dict: Evaluation metrics and their confidence intervals.
    """
    from sklearn.metrics import roc_curve, confusion_matrix, precision_score, accuracy_score, roc_auc_score, f1_score

    def find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity):
        fpr, tpr, thresholds = roc_curve(y_true, y_pred_prob)
        best_threshold = thresholds[np.argmin(np.abs(tpr - target_sensitivity))]

        ## -----------------------------------
        ## Find the threshold that gives a higher than target
        # best_threshold = None
        # for i, (threshold, current_tpr) in enumerate(zip(thresholds, tpr)):
        #     y_pred_at_threshold = (y_pred_prob >= threshold).astype(int)
        #     current_sensitivity = recall_score(y_true, y_pred_at_threshold)
        #
        #     if current_sensitivity >= target_sensitivity:
        #         best_threshold = threshold
        #         break


        return best_threshold



    def find_threshold_by_youden_index(y_true, y_pred_prob):
        fpr, tpr, thresholds = roc_curve(y_true, y_pred_prob)
        youden_index = tpr - fpr
        best_threshold = thresholds[np.argmax(youden_index)]
        return best_threshold

    def compute_confusion_metrics(y_true, y_pred):
        cm = confusion_matrix(y_true, y_pred)
        tn, fp, fn, tp = cm.ravel()
        return tn, fp, fn, tp

    # Determine threshold
    if predefined_thresh is not None:
        threshold = predefined_thresh
    elif target_sensitivity is not None:
        threshold = find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity)
    else:
        threshold = find_threshold_by_youden_index(y_true, y_pred_prob)

    # Predict using threshold
    y_pred = (y_pred_prob >= threshold).astype(int)
    tn, fp, fn, tp = compute_confusion_metrics(y_true, y_pred)

    # Compute metrics
    metrics = {
        'accuracy': accuracy_score(y_true, y_pred),
        'roc_auc': roc_auc_score(y_true, y_pred_prob),
        'specificity': tn / (tn + fp) if (tn + fp) else 0,
        'sensitivity': tp / (tp + fn) if (tp + fn) else 0,
        'ppv': precision_score(y_true, y_pred) if (tp + fp) else 0,
        'npv': tn / (tn + fn) if (tn + fn) else 0,
        'f1_score': f1_score(y_true, y_pred),
        'tn': tn, 'fp': fp, 'fn': fn, 'tp': tp,
        'threshold': threshold
    }

    # Optionally compute confidence intervals if function is defined
    ci = {}
    for k, v in metrics.items():
        if isinstance(v, (float, int)):
            ci[k] = compute_confidence_interval(v, len(y_true))

    return metrics, ci, y_pred


def compute_metrics1(y_true, y_pred, y_pred_prob, target_sensitivity=0.8):
    """
    Compute evaluation metrics and their confidence intervals,
    with precise sensitivity targeting.

    Parameters:
    y_true (array-like): True labels.
    y_pred (array-like): Predicted labels.
    y_pred_prob (array-like): Predicted probabilities.
    target_sensitivity (float, optional): Target sensitivity value.

    Returns:
    dict: Evaluation metrics and their confidence intervals.
    """
    from sklearn.metrics import roc_curve, confusion_matrix, precision_score, accuracy_score, roc_auc_score, f1_score

    # Find the threshold that most precisely matches target sensitivity
    def find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity):
        fpr, tpr, thresholds = roc_curve(y_true, y_pred_prob)

        # Find the threshold that gives sensitivity closest to target
        best_threshold = None
        # min_diff = float('inf')
        # for i, (threshold, current_tpr) in enumerate(zip(thresholds, tpr)):
        #     y_pred_at_threshold = (y_pred_prob >= threshold).astype(int)
        #     current_sensitivity = recall_score(y_true, y_pred_at_threshold)
        #
        #     diff = abs(current_sensitivity - target_sensitivity)
        #     if diff < min_diff:
        #         min_diff = diff
        #         best_threshold = threshold

        for i, (threshold, current_tpr) in enumerate(zip(thresholds, tpr)):
            y_pred_at_threshold = (y_pred_prob >= threshold).astype(int)
            current_sensitivity = recall_score(y_true, y_pred_at_threshold)

            if current_sensitivity >= target_sensitivity:
                best_threshold = threshold
                break

        print(f"best_threshold is {best_threshold}")
        return best_threshold

    # Find the optimal threshold
    threshold = find_threshold_for_sensitivity(y_true, y_pred_prob, target_sensitivity)

    # Use the found threshold for predictions
    y_pred_optimal = (y_pred_prob >= threshold).astype(int)
    print(f"y_pred_optimal is {y_pred_optimal}")

    # Compute metrics using optimal threshold predictions
    cm = confusion_matrix(y_true, y_pred_optimal)
    tn, fp, fn, tp = cm.ravel()

    print(f"tn is {tn}")
    print(f"fp is {fp}")
    print(f"fn is {fn}")
    print(f"tp is {tp}")
    print("============================================")

    metrics = {
        'accuracy': accuracy_score(y_true, y_pred_optimal),
        'roc_auc': roc_auc_score(y_true, y_pred_prob) if y_pred_prob is not None else None,
        'specificity': tn / (tn + fp) if (tn + fp) else 0,
        'sensitivity': tp / (tp + fn) if (tp + fn) else 0,
        'ppv': precision_score(y_true, y_pred_optimal),
        'npv': tn / (tn + fn) if (tn + fn) else 0,
        'f1_score': f1_score(y_true, y_pred_optimal)
    }

    ci = {}
    for metric, value in metrics.items():
        if value is not None:
            ci[metric] = compute_confidence_interval(value, y_true.size)

    return metrics, ci, y_pred_optimal


def compute_confidence_interval(metric_value, n, z=1.96):
    """
    Computes the confidence interval for a given metric value.
    """
    se = np.sqrt((metric_value * (1 - metric_value)) / n)
    ci_lower = metric_value - z * se
    ci_upper = metric_value + z * se
    return (ci_lower, ci_upper)


def calculate_confidence_interval_tpr(fpr, tpr, num_bootstraps=1000, random_state=42):
    rng = np.random.RandomState(random_state)
    bootstrapped_tprs = []

    for i in range(num_bootstraps):
        # Resample with replacement
        indices = rng.choice(len(fpr), size=len(fpr), replace=True)
        if len(np.unique(indices)) < 2:
            continue
        resampled_fpr = fpr[indices]
        resampled_tpr = tpr[indices]

        # Interpolate the resampled TPR at fixed FPR points
        interp_tpr = np.interp(fpr, np.sort(resampled_fpr), resampled_tpr[np.argsort(resampled_fpr)])
        bootstrapped_tprs.append(interp_tpr)

    # Calculate percentiles for confidence interval
    bootstrapped_tprs = np.array(bootstrapped_tprs)
    tpr_lower = np.percentile(bootstrapped_tprs, 2.5, axis=0)
    tpr_upper = np.percentile(bootstrapped_tprs, 97.5, axis=0)

    return tpr_lower, tpr_upper


def hyperparameter_tuning(clf, param_grid, X_train, y_train, name):
    """
    Perform hyperparameter tuning using GridSearchCV.

    Parameters:
    clf: The classifier to tune.
    param_grid (dict): The parameter grid to search over.
    X_train (pd.DataFrame): The training feature matrix.
    y_train (pd.Series): The training target vector.
    name (str): The name of the classifier.

    Returns:
    The best estimator found by GridSearchCV.
    """
    skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=17)


    grid_search = GridSearchCV(estimator=clf, param_grid=param_grid, cv=skf, n_jobs=-1, scoring='roc_auc')
    #print("grid_search done")
    grid_search.fit(X_train, y_train)
    #print("grid_search fit done")
    #print("Best parameters found by GridSearchCV:")
    #print("best_params: ", grid_search.best_params_)

    # Save grid search results
    #results_df = pd.DataFrame(grid_search.cv_results_)
    #results_df.to_csv(f'{name}_grid_search_results.csv', index=False)

    #print(f"Best parameters for {name}: {grid_search.best_params_}")
    #print(f"Best cross-validation score for {name}: {grid_search.best_score_}")

    return grid_search.best_estimator_

    #-----------------------------------------

    # random_search = RandomizedSearchCV(estimator=clf, param_distributions=param_grid, n_iter=20, cv=skf, n_jobs=-1,
    #                                    scoring='roc_auc', random_state=42)
    # print("random_search done")
    # random_search.fit(X_train, y_train)
    # print("grid_search fit done")
    #
    # return random_search.best_estimator_


def fill_na(X_train: pd.DataFrame,
           X_test: pd.DataFrame) -> (pd.DataFrame, pd.DataFrame):

    # Engineering missing values in numerical variables
    # numerical_cols = [col for col in X_train.columns if X_train[col].dtypes != '0']
    # numerical_cols = [col for col in X_train.columns if col not in categorical_columns]
    numerical_cols = X_train.columns

    for df1 in [X_train, X_test]:
        for col in numerical_cols:
            col_median = X_train[col].median()
            df1[col].fillna(col_median, inplace=True)

    # # Engineering missing values in categorical variables
    # for df2 in [X_train, X_test]:
    #     for col in categorical_columns:
    #         col_mod = X_train[col].mode()[0]
    #         df2[col].fillna(col_mod, inplace=True)

    return X_train, X_test



def remove_outliers(df):
    out_df = df.copy()

    # IQR
    for col in df.columns:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR
        out_df[col] = np.where(out_df[col] < lower, lower, out_df[col])
        out_df[col] = np.where(out_df[col] > upper, upper, out_df[col])

    return out_df




def normalize(X: pd.DataFrame):
    #scaler = MinMaxScaler()
    scaler = StandardScaler()
    cols = X.columns
    X = scaler.fit_transform(X)
    X = pd.DataFrame(X, columns=cols)
    return X



def train_test_split_evaluation(X, y,
                                test_size=0.3,
                                random_state=42,
                                tuning=False,
                                result_path="./results",
                                num_features=10,
                                resampling_method=None):


    # Split data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state)


    exclude_columns = ["Case"]
    X_train_Cases = X_train.loc[:, "Case"]
    X_test_Cases = X_test.loc[:, "Case"]

    X_train = X_train.loc[:, ~X.columns.isin(exclude_columns)]
    X_test = X_test.loc[:, ~X.columns.isin(exclude_columns)]



    # Fill Null values, remove outliers, and normalize
    X_train, X_test = fill_na(X_train, X_test)
    X_train = remove_outliers(X_train)
    X_test = remove_outliers(X_test)
    X_train = normalize(X_train)
    X_test = normalize(X_test)

    classifiers = get_classifiers()
    results = {}
    train_results = {}
    test_results = {}

    # Create directories
    roc_data_path = os.path.join(result_path, "ROC_data")
    ensure_directory_exists(roc_data_path)
    prob_data_path = os.path.join(result_path, "Prob_data")
    ensure_directory_exists(prob_data_path)
    roc_path = os.path.join(result_path, "ROC_curves")
    ensure_directory_exists(roc_path)
    waterfall_path = os.path.join(result_path, "Waterfall_plots")
    ensure_directory_exists(waterfall_path)
    calibration_path = os.path.join(result_path, "Calibration_plots")
    dca_path = os.path.join(result_path, "DCA_curves")
    shap_path = os.path.join(result_path, "Shapley_plots")
    importance_path = os.path.join(result_path, "Feature_Importance_plots")
    model_path = os.path.join(result_path, "Saved_Models")
    ensure_directory_exists(model_path)
    umap_path = os.path.join(result_path, "UMAP_plots")
    ensure_directory_exists(umap_path)
    other_plots_path = os.path.join(result_path, "Other_plots")
    ensure_directory_exists(other_plots_path)



    plt.figure(figsize=(10, 8))
    plt.title('ROC Curves', fontsize=20)
    plt.plot([0, 1], [0, 1], color='navy', linestyle='--', lw=2)

    for classifier_name, (clf, param_grid) in classifiers.items():
        if tuning:
            print(f"Hyperparameter tuning for {classifier_name} classifier")
            clf = hyperparameter_tuning(clf, param_grid, X_train, y_train, classifier_name)

        if resampling_method:
            X_train, y_train = resampling_method.fit_resample(X_train, y_train)


        #-----------------------------
        # Cross-validation evaluation for AUC_CV
        auc_cv_scores = cross_val_score(clf, X_train, y_train, cv=5, scoring='roc_auc')
        auc_cv_mean = auc_cv_scores.mean()
        print(f"Cross-validated AUC for {classifier_name}: {auc_cv_mean:.4f}")
        #-----------------------------

        #clf = CalibratedClassifierCV(estimator=RandomForestClassifier(), method='sigmoid', cv=5)
        clf.fit(X_train, y_train)

        # Train set evaluation
        y_pred_train = clf.predict(X_train)
        y_pred_prob_train = clf.predict_proba(X_train)[:, 1] if hasattr(clf, "predict_proba") else None
        metrics_train, ci_train, y_pred_train_optimal = compute_metrics(y_train, y_pred_prob_train, target_sensitivity=0.8)

        train_results[classifier_name] = {
            'metrics': metrics_train,
            'confidence_intervals': ci_train,
            'auc_cv': auc_cv_mean
        }

        # Compute ROC and save train data
        fpr_train, tpr_train, _ = roc_curve(y_train, y_pred_prob_train)
        tpr_lower_train, tpr_upper_train = calculate_confidence_interval_tpr(fpr_train, tpr_train)
        roc_auc_train = roc_auc_score(y_train, y_pred_prob_train)
        plot_roc_curve(fpr_train, tpr_train, roc_auc_train, tpr_lower_train, tpr_upper_train,
                       title=f'{classifier_name} Train ROC Curve',
                       filename=f'{classifier_name}_train_roc_curve_{num_features}_features.png',
                       output_dir=os.path.join(roc_path, classifier_name))

        # Save ROC data for train set
        roc_data_train_df = pd.DataFrame({
            'fpr': fpr_train,
            'tpr': tpr_train,
            'tpr_lower': tpr_lower_train,
            'tpr_upper': tpr_upper_train
        })
        roc_data_train_df.to_excel(os.path.join(roc_data_path, f'{classifier_name}_train_roc_data_{num_features}_features.xlsx'), index=False)

        # Save y_true and y_pred_prob for training set
        train_probs_df = pd.DataFrame({'y_true': y_train, 'y_pred_prob': y_pred_prob_train})
        train_probs_df.to_excel(
            os.path.join(prob_data_path, f'{classifier_name}_train_probs_{num_features}_features.xlsx'), index=False)

        # Test set evaluation
        y_pred_test = clf.predict(X_test)
        y_pred_prob_test = clf.predict_proba(X_test)[:, 1] if hasattr(clf, "predict_proba") else None

        print("*************")
        print(f"This is where the test metrics are going to be calculated for {classifier_name}")
        print("*************")
        metrics_test, ci_test, y_pred_test_optimal = compute_metrics(y_test, y_pred_prob_test, metrics_train.get('threshold', 'N/A'))
        #metrics_test, ci_test, y_pred_test_optimal = compute_metrics(y_test, y_pred_prob_test, target_sensitivity=0.8)
        test_results[classifier_name] = {
            'metrics': metrics_test,
            'confidence_intervals': ci_test
        }

        # Compute ROC and save test data
        fpr_test, tpr_test, _ = roc_curve(y_test, y_pred_prob_test)
        tpr_lower_test, tpr_upper_test = calculate_confidence_interval_tpr(fpr_test, tpr_test)
        roc_auc_test = roc_auc_score(y_test, y_pred_prob_test)
        plot_roc_curve(fpr_test, tpr_test, roc_auc_test, tpr_lower_test, tpr_upper_test,
                       title=f'{classifier_name} Test ROC Curve',
                       filename=f'{classifier_name}_test_roc_curve_{num_features}_features.png',
                       output_dir=os.path.join(roc_path, classifier_name))

        # Save ROC data for test set
        roc_data_test_df = pd.DataFrame({
            'fpr': fpr_test,
            'tpr': tpr_test,
            'tpr_lower': tpr_lower_test,
            'tpr_upper': tpr_upper_test
        })
        roc_data_test_df.to_excel(os.path.join(roc_data_path, f'{classifier_name}_test_roc_data_{num_features}_features.xlsx'), index=False)

        # Save y_true and y_pred_prob for test set
        test_probs_df = pd.DataFrame({'y_true': y_test, 'y_pred_prob': y_pred_prob_test})
        test_probs_df.to_excel(
            os.path.join(prob_data_path, f'{classifier_name}_test_probs_{num_features}_features.xlsx'), index=False)

        # Plot calibration, DCA, and Shapley values
        plot_calibration_curve(y_test, y_pred_prob_test, classifier_name, num_features, output_dir=calibration_path)
        plot_dca(y_test, y_pred_prob_test, classifier_name, num_features, output_dir=dca_path)
        # if hasattr(clf, 'feature_importances_'):
        #     plot_feature_importance(clf.feature_importances_, X.columns, 'Feature Importance',
        #                             f'{classifier_name}_featureImportance_{num_features}_features.png', output_dir=importance_path)
        if num_features > 1 and classifier_name in ['RandomForest', 'XGBoost', 'LightGBM']:
            plot_shap_values(clf, X_test, X.columns, f'{classifier_name}_shap_values_{num_features}_features.png', output_dir=shap_path)

        # Plot individual ROC for test set in aggregate
        plt.plot(fpr_test, tpr_test, lw=2, label=f'{classifier_name} (AUC = {roc_auc_test:.2f})')

        # # # Save final model
        # final_model = clf.fit(X, y)
        # final_model.feature_names = X.columns
        # final_model_filename = os.path.join(model_path, f'{classifier_name}_{num_features}_features.pkl')
        # joblib.dump(final_model, final_model_filename)

        # --------------------------------------
        # plottings
        # Plot feature importance for tree-based models
        feature_columns = X_train.columns  # Save feature names before normalization
        # if hasattr(clf, 'feature_importances_'):
        #     plot_feature_importance(clf.feature_importances_, feature_columns, 'Feature Importance',
        #                             f'{classifier_name}_featureImportance_{num_features}_features.png',
        #                             output_dir=importance_path)

        # Plot Shapley values
        # if num_features > 1 and name in ['RandomForest', 'XGBoost', 'LightGBM']:
        if num_features > 1:
            plot_shap_values(clf, X_train, feature_columns, f'{classifier_name}_shap_values_{num_features}_features.png',
                             output_dir=shap_path)


        plot_waterfall(y_train, y_pred_prob_train, classifier_name, num_features, sheet="",
                       pos_label="MMRd", neg_label="MMRp",
                       output_dir=os.path.join(waterfall_path, "training"), threshold=metrics_train.get('threshold', 'N/A'))

        train_df = pd.concat([X_train, y_train], axis=1)
        plot_umap(train_df, classifier_name, num_features, sheet="", features=feature_columns,
                  exclude_cols=exclude_columns, outcome_col='Outcome',
                  title='UMAP Projection of Radiomics Features for Training Cohort', filepath=os.path.join(umap_path, "training"))

        # plot_umap_tsne_3d(train_df, classifier_name, num_features, sheet="", features=feature_columns,
        #                   exclude_cols=exclude_columns, outcome_col='Outcome',
        #                   title='UMAP Projection of Radiomics Features for Training Cohort', filepath=os.path.join(umap_path, "training"))

        #----------
        plot_waterfall(y_test, y_pred_prob_test, classifier_name, num_features, sheet="",
                       pos_label="MMRd", neg_label="MMRp",
                       output_dir=os.path.join(waterfall_path, "test"), threshold=metrics_test.get('threshold', 'N/A'))

        test_df = pd.concat([X_test, y_test], axis=1)
        plot_umap(test_df, classifier_name, num_features, sheet="", features=feature_columns,
                  exclude_cols=exclude_columns, outcome_col='Outcome',
                  title='UMAP Projection of Radiomics Features for Test Cohort',
                  filepath=os.path.join(umap_path, "test"))

        # plot_umap_tsne_3d(test_df, classifier_name, num_features, sheet="", features=feature_columns,
        #                   exclude_cols=exclude_columns, outcome_col='Outcome',
        #                   title='UMAP Projection of Radiomics Features for Test Cohort',
        #                   filepath=os.path.join(umap_path, "test"))



        thresholds = np.linspace(0, 1, 100)
        train_recalls = [recall_score(y_train, y_pred_prob_train >= t) for t in thresholds]
        test_recalls = [recall_score(y_test, y_pred_prob_test >= t) for t in thresholds]

        selected_threshold = metrics_test.get('threshold', None)

        plt.figure()
        plt.plot(thresholds, train_recalls, label=f"Train Sensitivity ({metrics_train.get('sensitivity', None):.2f})")
        plt.plot(thresholds, test_recalls, label=f"Train Sensitivity ({metrics_test.get('sensitivity', None):.2f})")
        plt.axhline(0.8, color='red', linestyle='--', label='Target Sensitivity (0.8)')

        if selected_threshold is not None:
            plt.axvline(selected_threshold, color='purple', linestyle='--',
                        label=f'Selected Threshold = {selected_threshold:.2f}')

        plt.xlabel('Threshold')
        plt.ylabel('Sensitivity')
        plt.legend()
        filepath1 = os.path.join(other_plots_path, f'SensitivityVSThreshold_{num_features}_features.png')
        plt.savefig(filepath1, dpi=300)
        plt.close()

        # --------------------------------------

        # #------------------------------------------------------
        # # Save case_id, predicted outcome, actual outcome, and probability scores
        # output_df_train = pd.DataFrame({
        #     'Case_ID': X_train_Cases,  # Replace 'Case_ID' if column name differs
        #     'Probability_Score': y_pred_prob_train,
        #     'Predicted_Outcome': y_pred_train_optimal,  # Predicted outcome
        #     'Actual_Outcome': y_train  # Actual outcome
        # })
        #
        # output_excel_path_train = os.path.join(prob_data_path,
        #                                       f"case_predictions_train_{classifier_name}_{num_features}.xlsx")
        # save_excel_sheet(output_df_train, output_excel_path_train, str(num_features), False)
        # # ------------------------------------------------------
        # # Save case_id, predicted outcome, actual outcome, and probability scores
        # output_df_test = pd.DataFrame({
        #     'Case_ID': X_test_Cases,  # Replace 'Case_ID' if column name differs
        #     'Probability_Score': y_pred_prob_test,
        #     'Predicted_Outcome': y_pred_test_optimal,  # Predicted outcome
        #     'Actual_Outcome': y_test  # Actual outcome
        # })
        #
        # output_excel_path_test = os.path.join(prob_data_path,
        #                                        f"case_predictions_test_{classifier_name}_{num_features}.xlsx")
        # save_excel_sheet(output_df_test, output_excel_path_test, str(num_features), False)
        # # ------------------------------------------------------

    # Finalize and save the aggregated ROC plot
    plt.xlabel('False Positive Rate', fontsize=18)
    plt.ylabel('True Positive Rate', fontsize=18)
    plt.legend(loc='lower right', fontsize=16)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    filepath = os.path.join(roc_path, f'roc_curve_{num_features}_features.png')
    plt.savefig(filepath, dpi=300)

    results['train'] = train_results
    results['test'] = test_results

    return results


def cross_validation_evaluation(X, y, cv_folds=5, tuning=False, result_path="./results",
                                num_features=10, resampling_method=None):
    classifiers = get_classifiers()
    results = {}

    exclude_columns = ["Patient_ID"]
    X_Cases = X.loc[:, "Patient_ID"]
    X = X.loc[:, ~X.columns.isin(exclude_columns)]

    roc_data_path = os.path.join(result_path, "ROC_data")
    ensure_directory_exists(roc_data_path)
    prob_data_path = os.path.join(result_path, "Prob_data")
    ensure_directory_exists(prob_data_path)
    roc_path = os.path.join(result_path, "ROC_curves")
    ensure_directory_exists(roc_path)
    waterfall_path = os.path.join(result_path, "Waterfall_plots", "Training")
    ensure_directory_exists(waterfall_path)
    calibration_path = os.path.join(result_path, "Calibration_plots")
    dca_path = os.path.join(result_path, "DCA_curves")
    shap_path = os.path.join(result_path, "Shapley_plots")
    importance_path = os.path.join(result_path, "Feature_Importance_plots")
    model_path = os.path.join(result_path, "Saved_Models")
    ensure_directory_exists(model_path)
    umap_path = os.path.join(result_path, "UMAP_plots", "Training")
    ensure_directory_exists(umap_path)

    plt.figure(figsize=(10, 8))
    plt.title('Averaged ROC Curves', fontsize=20)
    plt.plot([0, 1], [0, 1], color='navy', linestyle='--', lw=2)

    skf = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=17)

    for name, (clf, param_grid) in classifiers.items():
        metrics_list = []
        fpr_list = []
        tpr_list = []
        thresholds_list = []
        auc_list = []
        y_true_all_folds = []
        y_pred_prob_all_folds = []


        if tuning:
            print(f"Hyperparameter tuning for {name} classifier")
            clf = hyperparameter_tuning(clf, param_grid, X, y, name)

        for fold, (train_index, test_index) in enumerate(skf.split(X, y), 1):
            X_train, X_test = X.iloc[train_index], X.iloc[test_index]
            y_train, y_test = y.iloc[train_index], y.iloc[test_index]

            # # Preprocessing
            X_train = remove_outliers(X_train)
            X_test = remove_outliers(X_test)
            # X_train = normalize(X_train)
            # X_test = normalize(X_test)
            X_train, X_test = fill_na(X_train, X_test)
            if resampling_method:
                X_train, y_train = resampling_method.fit_resample(X_train, y_train)


            clf.fit(X_train, y_train)

            y_pred = clf.predict(X_test)
            y_pred_prob = clf.predict_proba(X_test)[:, 1] if hasattr(clf, "predict_proba") else None

            y_true_all_folds.extend(y_test.tolist())
            y_pred_prob_all_folds.extend(y_pred_prob.tolist() if y_pred_prob is not None else [])


            ## ---------------------------------------
            # # Plot feature importance for tree-based models
            # feature_columns = X_train.columns  # Save feature names before normalization
            # if hasattr(clf, 'feature_importances_'):
            #     plot_feature_importance(clf.feature_importances_, feature_columns, 'Feature Importance',
            #                             f'{name}_featureImportance_{num_features}_features.png', output_dir=importance_path)
            #
            # # Plot Shapley values
            # #if num_features > 1 and name in ['RandomForest', 'XGBoost', 'LightGBM']:
            # if num_features > 1:
            #     plot_shap_values(clf, X_test, feature_columns, f'{name}_shap_values_{num_features}_features.png', output_dir=shap_path)
            ## ---------------------------------------

        # Collect data for ROC plotting
        y_true_train = pd.Series(y_true_all_folds)
        y_pred_prob_train = pd.Series(y_pred_prob_all_folds)


        fpr, tpr, thresholds = roc_curve(y_true_train, y_pred_prob_train)
        #roc_auc = roc_auc_score(y_true_train, y_pred_prob_train)
        roc_auc = auc(fpr, tpr)

        #tpr_lower, tpr_upper = calculate_confidence_interval_tpr(fpr, tpr)

        # Confidence interval calculation with stratified bootstrap resampling
        bootstrapped_tpr = []
        n_bootstraps = 1000  # Increase the number of bootstraps for more stability
        rng_seed = 42  # Random seed for reproducibility
        rng = np.random.RandomState(rng_seed)

        for i in range(n_bootstraps):
            # Stratified bootstrapping to maintain class distribution
            indices = rng.choice(np.arange(len(y_true_train)), size=len(y_true_train),
                                 replace=True)
            y_sample = y_true_train.iloc[indices]
            if len(np.unique(y_sample)) < 2:
                # We need at least one positive and one negative sample to compute ROC
                continue
            fpr_boot, tpr_boot, _ = roc_curve(y_sample, y_pred_prob_train[indices])
            bootstrapped_tpr.append(
                np.interp(fpr, fpr_boot, tpr_boot))  # Interpolation to align FPR

        bootstrapped_tpr = np.array(bootstrapped_tpr)
        tpr_lower = np.percentile(bootstrapped_tpr, 2.5, axis=0)
        tpr_upper = np.percentile(bootstrapped_tpr, 97.5, axis=0)







        # Create a DataFrame with TPR, FPR, AUC, and confidence intervals
        roc_df = pd.DataFrame({
            'fpr': fpr,
            'tpr': tpr,
            'thresholds': thresholds,
            'tpr_lower': tpr_lower,
            'tpr_upper': tpr_upper,
        })

        # Add the AUC value to a separate column (you can also store it separately if needed)
        roc_df['AUC'] = roc_auc

        roc_df.to_excel(
            os.path.join(roc_data_path, f'{name}_averaged_roc_data_{num_features}_features.xlsx'), index=False)

        # --------------------------------------






        # Compute metrics after all folds
        metrics, ci, y_pred_optimal = compute_metrics(pd.Series(y_true_all_folds), pd.Series(y_pred_prob_all_folds), target_sensitivity=0.8)
        averaged_metrics = {metric: metrics[metric] for metric in metrics}
        #ci = {metric: compute_confidence_interval(averaged_metrics[metric], y.size) for metric in averaged_metrics}

        results[name] = {
            'metrics': averaged_metrics,
            'confidence_intervals': ci
        }



        # Save predicted probabilities and true labels
        prob_df = pd.DataFrame({'y_true': y_true_all_folds, 'y_pred_prob': y_pred_prob_all_folds})
        prob_df.to_excel(os.path.join(prob_data_path, f'{name}_predicted_probs_{num_features}_features.xlsx'), index=False)




        # save the final model
        selected_thresh = metrics.get('target_threshold', '0.5')
        if isinstance(selected_thresh, str):  # Convert to float if possible
            try:
                selected_thresh = float(selected_thresh)
            except ValueError:
                selected_thresh = 0.5

        # Train the final model on the entire dataset and save it
        final_model = clf.fit(X, y)
        final_model.feature_names = X.columns
        final_model.selected_thresh = selected_thresh
        final_model_filename = os.path.join(model_path, f'{name}_{num_features}_features.pkl')
        joblib.dump(final_model, final_model_filename)

        # Save feature names to a text file
        feature_names_filename = os.path.splitext(final_model_filename)[0] + '_features.txt'
        with open(feature_names_filename, 'w') as f:
            for feature in final_model.feature_names:
                f.write(f"{feature}\n")

        # --------------------------------------
        # plottings
        if SAVE_PLOTS:
            # # Plot feature importance for tree-based models
            feature_columns = X_train.columns  # Save feature names before normalization
            # if hasattr(clf, 'feature_importances_'):
            #     plot_feature_importance(clf.feature_importances_, feature_columns, 'Feature Importance',
            #                             f'{name}_featureImportance_{num_features}_features.png',
            #                             output_dir=importance_path)

            # Plot Shapley values
            # if num_features > 1 and name in ['RandomForest', 'XGBoost', 'LightGBM']:
            if num_features > 1:
                plot_shap_values(clf, X, feature_columns, f'{name}_shap_values_{num_features}_features.png', output_dir=shap_path)

            plot_waterfall(y_true_train, y_pred_prob_train, name, num_features, sheet="",
                           pos_label="PNET grade high", neg_label="PNET grade low",
                           output_dir=waterfall_path, threshold=metrics.get('target_threshold', 'N/A'))




            # train_df = pd.concat([X, y], axis = 1)
            # plot_umap(train_df, name, num_features, sheet="", features=feature_columns,
            #           exclude_cols=exclude_columns + ['Size'], outcome_col='Grade',
            #           title='UMAP Projection of Radiomics Features for Training Cohort', filepath=umap_path)
            #
            # plot_umap_tsne_3d(train_df, name, num_features, sheet="", features=feature_columns,
            #           exclude_cols=exclude_columns + ['Size'], outcome_col='Grade',
            #           title='UMAP Projection of Radiomics Features for Training Cohort', filepath=umap_path)

        # --------------------------------------
        print(f"Testing trained {name} on train data with {num_features} features...")

        y_pred_prob_train = clf.predict_proba(X)[:, 1] if hasattr(clf, "predict_proba") else None

        # Save case_id, predicted outcome, actual outcome, and probability scores
        output_df_train = pd.DataFrame({
            'Patient_ID': X_Cases,  # Replace 'Case_ID' if column name differs
            'Probability_Score': y_pred_prob_train,
            'Predicted_Outcome': y_pred_optimal,  # Predicted outcome
            'Actual_Outcome': y  # Actual outcome
        })

        output_excel_path_train = os.path.join(prob_data_path,
                                               f"case_predictions_train_{name}_{num_features}.xlsx")
        save_excel_sheet(output_df_train, output_excel_path_train, str(num_features), False)

        cm_matrix = pd.DataFrame({
            'TP': [f"{metrics.get('tp', 'N/A')}"],
            'TN': [f"{metrics.get('tn', 'N/A')}"],
            'FP': [f"{metrics.get('fp', 'N/A')}"],
            'FN': [f"{metrics.get('fn', 'N/A')}"],
            'Threshold': [f"{metrics.get('target_threshold', 'N/A'):.8f}"],
        })

        save_excel_sheet(cm_matrix, output_excel_path_train, "cm_" + str(num_features), False)

        # --------------------------------------


        # #-------------------------------
        # # Extract and save feature weights and bias for supported models
        # model = final_model
        # if isinstance(final_model, Pipeline):  # If the model is in a Pipeline
        #     model = final_model.named_steps['lr']  # Extract the model step (assumes 'lr', adjust as needed)
        #
        # if hasattr(model, 'coef_'):  # For Logistic Regression and Linear SVM
        #     weights = model.coef_[0]  # Coefficients for the features
        #     bias = model.intercept_[0]  # Intercept term
        #
        #     # Save weights and bias to a CSV file
        #     feature_weights_df = pd.DataFrame({
        #         'Feature': final_model.feature_names,
        #         'Weight': weights
        #     })
        #     weights_filename = os.path.splitext(final_model_filename)[0] + '_feature_weights.csv'
        #     feature_weights_df.to_csv(weights_filename, index=False)
        #
        #     # Save bias to a text file
        #     bias_filename = os.path.splitext(final_model_filename)[0] + '_bias.txt'
        #     with open(bias_filename, 'w') as f:
        #         f.write(f"Bias (Intercept): {bias}\n")
        #
        # elif hasattr(model, 'feature_importances_'):  # For Random Forest and other tree-based models
        #     feature_importances = model.feature_importances_  # Feature importances
        #     feature_importances_df = pd.DataFrame({
        #         'Feature': final_model.feature_names,
        #         'Importance': feature_importances
        #     })
        #     importance_filename = os.path.splitext(final_model_filename)[0] + '_feature_importances.csv'
        #     feature_importances_df.to_csv(importance_filename, index=False)
        #
        # elif isinstance(model, GaussianNB):  # For Naive Bayes
        #     print("Naive Bayes does not provide feature weights; skipping this step.")
        #
        # else:
        #     print(f"Feature weight extraction is not supported for the model {type(model)}. Skipping.")
        # #-------------------------------

    if SAVE_PLOTS:
        plt.xlabel('False Positive Rate', fontsize=18)
        plt.ylabel('True Positive Rate', fontsize=18)
        plt.legend(loc='lower right', fontsize=16)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        filepath = os.path.join(roc_path, f'averaged_roc_curve_{num_features}_features.png')
        plt.savefig(filepath, dpi=600)

    return results





def train_test_split_evaluation_CV(X, y,
                                test_size=0.3,
                                random_state=42,
                                tuning=False,
                                result_path="./results",
                                num_features=10,
                                resampling_method=None):


    # Split data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state)


    exclude_columns = ["Case"]
    X_train_Cases = X_train.loc[:, "Case"]
    X_test_Cases = X_test.loc[:, "Case"]

    X_train = X_train.loc[:, ~X_train.columns.isin(exclude_columns)]
    X_test = X_test.loc[:, ~X_test.columns.isin(exclude_columns)]

    classifiers = get_classifiers()
    results = {}
    train_results = {}
    test_results = {}

    # Create directories
    roc_data_path = os.path.join(result_path, "ROC_data")
    ensure_directory_exists(roc_data_path)
    prob_data_path = os.path.join(result_path, "Prob_data")
    ensure_directory_exists(prob_data_path)
    roc_path = os.path.join(result_path, "ROC_curves")
    ensure_directory_exists(roc_path)
    waterfall_path = os.path.join(result_path, "Waterfall_plots")
    ensure_directory_exists(waterfall_path)
    calibration_path = os.path.join(result_path, "Calibration_plots")
    dca_path = os.path.join(result_path, "DCA_curves")
    shap_path = os.path.join(result_path, "Shapley_plots")
    importance_path = os.path.join(result_path, "Feature_Importance_plots")
    model_path = os.path.join(result_path, "Saved_Models")
    ensure_directory_exists(model_path)
    umap_path = os.path.join(result_path, "UMAP_plots")
    ensure_directory_exists(umap_path)
    other_plots_path = os.path.join(result_path, "Other_plots")
    ensure_directory_exists(other_plots_path)



    plt.figure(figsize=(10, 8))
    plt.title('ROC Curves', fontsize=20)
    plt.plot([0, 1], [0, 1], color='navy', linestyle='--', lw=2)

    skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=17)

    for classifier_name, (clf, param_grid) in classifiers.items():
        y_true_all_folds = []
        y_pred_prob_all_folds = []

        if tuning:
            print(f"Hyperparameter tuning for {classifier_name} classifier")
            clf = hyperparameter_tuning(clf, param_grid, X_train, y_train, classifier_name)

        for fold, (train_index, test_index) in enumerate(skf.split(X_train, y_train), 1):
            X_train_fold, X_test_fold = X_train.iloc[train_index], X_train.iloc[test_index]
            y_train_fold, y_test_fold = y_train.iloc[train_index], y_train.iloc[test_index]

            # Fill Null values, remove outliers, and normalize
            X_train_fold, X_test_fold = fill_na(X_train_fold, X_test_fold)
            X_train_fold = remove_outliers(X_train_fold)
            X_test_fold = remove_outliers(X_test_fold)
            X_train_fold = normalize(X_train_fold)
            X_test_fold = normalize(X_test_fold)

            if resampling_method:
                X_train_fold, y_train_fold = resampling_method.fit_resample(X_train_fold, y_train_fold)

            clf.fit(X_train_fold, y_train_fold)

            y_pred = clf.predict(X_test_fold)
            y_pred_prob = clf.predict_proba(X_test_fold)[:, 1] if hasattr(clf, "predict_proba") else None

            y_true_all_folds.extend(y_test_fold.tolist())
            y_pred_prob_all_folds.extend(y_pred_prob.tolist() if y_pred_prob is not None else [])



        y_true_train = pd.Series(y_true_all_folds)
        y_pred_prob_train = pd.Series(y_pred_prob_all_folds)


        #========================================================================================================

        # Train set evaluation
        metrics_train, ci_train, y_pred_train_optimal = compute_metrics(y_train, y_pred_prob_train, target_sensitivity=0.8)

        # Compute ROC and save train data
        fpr_train, tpr_train, _ = roc_curve(y_train, y_pred_prob_train)
        tpr_lower_train, tpr_upper_train = calculate_confidence_interval_tpr(fpr_train, tpr_train)
        roc_auc_train = roc_auc_score(y_train, y_pred_prob_train)

        train_results[classifier_name] = {
            'metrics': metrics_train,
            'confidence_intervals': ci_train,
            'auc': roc_auc_train
        }

        plot_roc_curve(fpr_train, tpr_train, roc_auc_train, tpr_lower_train, tpr_upper_train,
                       title=f'{classifier_name} Train ROC Curve',
                       filename=f'{classifier_name}_train_roc_curve_{num_features}_features.png',
                       output_dir=os.path.join(roc_path, classifier_name))

        # Save ROC data for train set
        roc_data_train_df = pd.DataFrame({
            'fpr': fpr_train,
            'tpr': tpr_train,
            'tpr_lower': tpr_lower_train,
            'tpr_upper': tpr_upper_train
        })
        roc_data_train_df.to_excel(os.path.join(roc_data_path, f'{classifier_name}_train_roc_data_{num_features}_features.xlsx'), index=False)

        # Save y_true and y_pred_prob for training set
        train_probs_df = pd.DataFrame({'y_true': y_train, 'y_pred_prob': y_pred_prob_train})
        train_probs_df.to_excel(
            os.path.join(prob_data_path, f'{classifier_name}_train_probs_{num_features}_features.xlsx'), index=False)

        # Test set evaluation
        y_pred_test = clf.predict(X_test)
        y_pred_prob_test = clf.predict_proba(X_test)[:, 1] if hasattr(clf, "predict_proba") else None

        print("*************")
        print(f"This is where the test metrics are going to be calculated for {classifier_name}")
        print("*************")
        #metrics_test, ci_test, y_pred_test_optimal = compute_metrics(y_test, y_pred_prob_test, predefined_thresh = metrics_train.get('threshold', 'N/A'))
        metrics_test, ci_test, y_pred_test_optimal = compute_metrics(y_test, y_pred_prob_test, target_sensitivity=0.8)
        test_results[classifier_name] = {
            'metrics': metrics_test,
            'confidence_intervals': ci_test
        }

        # Compute ROC and save test data
        fpr_test, tpr_test, _ = roc_curve(y_test, y_pred_prob_test)
        tpr_lower_test, tpr_upper_test = calculate_confidence_interval_tpr(fpr_test, tpr_test)
        roc_auc_test = roc_auc_score(y_test, y_pred_prob_test)
        plot_roc_curve(fpr_test, tpr_test, roc_auc_test, tpr_lower_test, tpr_upper_test,
                       title=f'{classifier_name} Test ROC Curve',
                       filename=f'{classifier_name}_test_roc_curve_{num_features}_features.png',
                       output_dir=os.path.join(roc_path, classifier_name))

        # Save ROC data for test set
        roc_data_test_df = pd.DataFrame({
            'fpr': fpr_test,
            'tpr': tpr_test,
            'tpr_lower': tpr_lower_test,
            'tpr_upper': tpr_upper_test
        })
        roc_data_test_df.to_excel(os.path.join(roc_data_path, f'{classifier_name}_test_roc_data_{num_features}_features.xlsx'), index=False)

        # Save y_true and y_pred_prob for test set
        test_probs_df = pd.DataFrame({'y_true': y_test, 'y_pred_prob': y_pred_prob_test})
        test_probs_df.to_excel(
            os.path.join(prob_data_path, f'{classifier_name}_test_probs_{num_features}_features.xlsx'), index=False)

        # # Plot calibration, DCA, and Shapley values
        # plot_calibration_curve(y_test, y_pred_prob_test, classifier_name, num_features, output_dir=calibration_path)
        # plot_dca(y_test, y_pred_prob_test, classifier_name, num_features, output_dir=dca_path)
        # if hasattr(clf, 'feature_importances_'):
        #     plot_feature_importance(clf.feature_importances_, X.columns, 'Feature Importance',
        #                             f'{classifier_name}_featureImportance_{num_features}_features.png', output_dir=importance_path)
        # if num_features > 1 and classifier_name in ['RandomForest', 'XGBoost', 'LightGBM']:
        #     plot_shap_values(clf, X_test, X.columns, f'{classifier_name}_shap_values_{num_features}_features.png', output_dir=shap_path)

        # Plot individual ROC for test set in aggregate
        plt.plot(fpr_test, tpr_test, lw=2, label=f'{classifier_name} (AUC = {roc_auc_test:.2f})')

        # # Save final model
        # final_model = clf.fit(X, y)
        # final_model.feature_names = X.columns
        # final_model_filename = os.path.join(model_path, f'{classifier_name}_{num_features}_features.pkl')
        # joblib.dump(final_model, final_model_filename)

        # --------------------------------------
        # plottings
        # Plot feature importance for tree-based models
        # feature_columns = X_train.columns  # Save feature names before normalization
        # if hasattr(clf, 'feature_importances_'):
        #     plot_feature_importance(clf.feature_importances_, feature_columns, 'Feature Importance',
        #                             f'{classifier_name}_featureImportance_{num_features}_features.png',
        #                             output_dir=importance_path)

        # # Plot Shapley values
        # # if num_features > 1 and name in ['RandomForest', 'XGBoost', 'LightGBM']:
        # if num_features > 1:
        #     plot_shap_values(clf, X_train, feature_columns, f'{classifier_name}_shap_values_{num_features}_features.png',
        #                      output_dir=shap_path)
        #
        #
        # plot_waterfall(y_train, y_pred_prob_train, classifier_name, num_features, sheet="",
        #                pos_label="MMRd", neg_label="MMRp",
        #                output_dir=os.path.join(waterfall_path, "training"), threshold=metrics_train.get('threshold', 'N/A'))

        train_df = pd.concat([X_train, y_train], axis=1)
        # plot_umap(train_df, classifier_name, num_features, sheet="", features=feature_columns,
        #           exclude_cols=exclude_columns, outcome_col='Outcome',
        #           title='UMAP Projection of Radiomics Features for Training Cohort', filepath=os.path.join(umap_path, "training"))

        # plot_umap_tsne_3d(train_df, classifier_name, num_features, sheet="", features=feature_columns,
        #                   exclude_cols=exclude_columns, outcome_col='Outcome',
        #                   title='UMAP Projection of Radiomics Features for Training Cohort', filepath=os.path.join(umap_path, "training"))

        #----------
        # plot_waterfall(y_test, y_pred_prob_test, classifier_name, num_features, sheet="",
        #                pos_label="MMRd", neg_label="MMRp",
        #                output_dir=os.path.join(waterfall_path, "test"), threshold=metrics_test.get('threshold', 'N/A'))

        test_df = pd.concat([X_test, y_test], axis=1)
        # plot_umap(test_df, classifier_name, num_features, sheet="", features=feature_columns,
        #           exclude_cols=exclude_columns, outcome_col='Outcome',
        #           title='UMAP Projection of Radiomics Features for Test Cohort',
        #           filepath=os.path.join(umap_path, "test"))

        # plot_umap_tsne_3d(test_df, classifier_name, num_features, sheet="", features=feature_columns,
        #                   exclude_cols=exclude_columns, outcome_col='Outcome',
        #                   title='UMAP Projection of Radiomics Features for Test Cohort',
        #                   filepath=os.path.join(umap_path, "test"))



        thresholds = np.linspace(0, 1, 100)
        train_recalls = [recall_score(y_train, y_pred_prob_train >= t) for t in thresholds]
        test_recalls = [recall_score(y_test, y_pred_prob_test >= t) for t in thresholds]

        selected_threshold = metrics_test.get('threshold', None)

        plt.figure()
        plt.plot(thresholds, train_recalls, label=f"Train Sensitivity ({metrics_train.get('sensitivity', None):.2f})")
        plt.plot(thresholds, test_recalls, label=f"Train Sensitivity ({metrics_test.get('sensitivity', None):.2f})")
        plt.axhline(0.8, color='red', linestyle='--', label='Target Sensitivity (0.8)')

        if selected_threshold is not None:
            plt.axvline(selected_threshold, color='purple', linestyle='--',
                        label=f'Selected Threshold = {selected_threshold:.2f}')

        plt.xlabel('Threshold')
        plt.ylabel('Sensitivity')
        plt.legend()
        filepath1 = os.path.join(other_plots_path, f'SensitivityVSThreshold_{num_features}_features.png')
        plt.savefig(filepath1, dpi=300)
        plt.close()

        # --------------------------------------

        #------------------------------------------------------
        # Save case_id, predicted outcome, actual outcome, and probability scores
        output_df_train = pd.DataFrame({
            'Case_ID': X_train_Cases,  # Replace 'Case_ID' if column name differs
            'Probability_Score': y_pred_prob_train,
            'Predicted_Outcome': y_pred_train_optimal,  # Predicted outcome
            'Actual_Outcome': y_train  # Actual outcome
        })

        output_excel_path_train = os.path.join(prob_data_path,
                                              f"case_predictions_train_{classifier_name}_{num_features}.xlsx")
        save_excel_sheet(output_df_train, output_excel_path_train, str(num_features), False)
        # ------------------------------------------------------
        # Save case_id, predicted outcome, actual outcome, and probability scores
        output_df_test = pd.DataFrame({
            'Case_ID': X_test_Cases,  # Replace 'Case_ID' if column name differs
            'Probability_Score': y_pred_prob_test,
            'Predicted_Outcome': y_pred_test_optimal,  # Predicted outcome
            'Actual_Outcome': y_test  # Actual outcome
        })

        output_excel_path_test = os.path.join(prob_data_path,
                                               f"case_predictions_test_{classifier_name}_{num_features}.xlsx")
        save_excel_sheet(output_df_test, output_excel_path_test, str(num_features), False)
        # ------------------------------------------------------

    # Finalize and save the aggregated ROC plot
    plt.xlabel('False Positive Rate', fontsize=18)
    plt.ylabel('True Positive Rate', fontsize=18)
    plt.legend(loc='lower right', fontsize=16)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    filepath = os.path.join(roc_path, f'roc_curve_{num_features}_features.png')
    plt.savefig(filepath, dpi=300)

    results['train'] = train_results
    results['test'] = test_results

    return results




def evaluate_models(X, y, method='train_test_split', **kwargs):
    """
    Evaluate models using the specified method.
    Parameters:
    X (pd.DataFrame): Feature matrix.
    y (pd.Series): Target vector.
    method (str): Evaluation method ('train_test_split' or 'cross_validation').
    kwargs: Additional arguments for the evaluation methods.
    Returns:
    dict: Evaluation results for each classifier.
    """
    if method == 'train_test_split':
        return train_test_split_evaluation(X, y, **kwargs)
        #return train_test_split_evaluation_CV(X, y, **kwargs)
    elif method == 'cross_validation':
        return cross_validation_evaluation(X, y, **kwargs)
    elif method == 'cv_feature_selection_model_building':
        return cv_feature_selection_model_building_evaluation(X, y, **kwargs)
    else:
        raise ValueError("Invalid method. Choose 'train_test_split' or 'cross_validation'.")


def ensure_directory_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


# Function to plot ROC curve
def plot_roc_curve(fpr, tpr, auc, tpr_lower, tpr_upper, title, filename, output_dir='./plots'):
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, filename)

    fpr = [0] + fpr
    tpr = [0] + tpr
    tpr_lower = [0] + tpr_lower
    tpr_upper = [0] + tpr_upper

    plt.figure()
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (area = {auc:.2f})')
    plt.fill_between(fpr, tpr_lower, tpr_upper, color = 'darkorange', alpha = 0.2)
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate', fontsize=18)
    plt.ylabel('True Positive Rate', fontsize=18)
    plt.title(title, fontsize=20)
    plt.legend(loc="lower right", fontsize=16)
    plt.savefig(filepath)
    plt.close()





def plot_calibration_curve(y_true, y_pred_prob, classifier_name, num_features, num_bins=10, output_dir='./plots'):
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, f'calibration_curve_{classifier_name}_{num_features}_features.png')

    plt.figure()
    prob_true, prob_pred = calibration_curve(y_true, y_pred_prob, n_bins=num_bins)
    plt.plot(prob_pred, prob_true, marker='o', label=f'{classifier_name}')
    plt.plot([0, 1], [0, 1], linestyle='--', label='Perfectly Calibrated')
    plt.xlabel('Predicted Probability')
    plt.ylabel('True Probability')
    plt.title(f'Calibration Curve - {classifier_name}')
    plt.legend(loc='upper left', fontsize=16)
    plt.savefig(filepath)
    plt.close()


def plot_dca(y_true, y_pred_prob, classifier_name, num_features, output_dir='./plots'):
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, f'dca_{classifier_name}_{num_features}_features.png')

    thresholds = np.linspace(0.1, 0.85, 50)
    net_benefits = []
    treat_all = []

    n_patients = len(y_true)
    n_positive = np.sum(y_true)

    for threshold in thresholds:
        # Model strategy
        y_pred = (y_pred_prob >= threshold).astype(int)
        tp = np.sum((y_pred == 1) & (y_true == 1))
        fp = np.sum((y_pred == 1) & (y_true == 0))

        # Net benefit calculations
        epsilon = 1e-6
        net_benefit = (tp / n_patients) - (fp / n_patients) * (threshold / max(epsilon, 1 - threshold))
        treat_all_benefit = (n_positive / n_patients) - ((n_patients - n_positive) / n_patients) * (threshold / max(epsilon, 1 - threshold))
        treat_all_benefit = max(treat_all_benefit, -1)  # Cap extreme negative values

        net_benefits.append(net_benefit)
        treat_all.append(treat_all_benefit)

    net_benefits = gaussian_filter1d(net_benefits, sigma=1)

    # Plot
    plt.figure(figsize=(8, 6))
    plt.plot(thresholds, net_benefits, label='Model', color='blue', linewidth=2)
    plt.plot(thresholds, treat_all, label='Treat all', color='black', linestyle='-', linewidth=2)
    plt.axhline(y=0, color='gray', linestyle='--', label='Treat none', linewidth=2)
    plt.fill_between(thresholds, 0, net_benefits, where=(np.array(net_benefits) > 0), color='lightblue', alpha=0.3)

    plt.xlabel('Threshold Probability', fontsize=14)
    plt.ylabel('Net Benefit', fontsize=14)
    plt.title(f'Decision Curve Analysis - {classifier_name}', fontsize=16)
    plt.legend(loc='upper right', fontsize=12)
    plt.grid(True, linestyle='--', alpha=0.6)
    plt.savefig(filepath)
    plt.close()






# Function to plot confusion matrix
def plot_confusion_matrix(y_true, y_pred, title, filename, output_dir='./plots'):
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, filename)

    cm = confusion_matrix(y_true, y_pred)
    plt.figure(figsize=(10, 7))
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues')
    plt.xlabel('Predicted')
    plt.ylabel('True')
    plt.title(title)
    plt.savefig(filepath)
    plt.close()


# Function to plot precision-recall curve
def plot_precision_recall_curve(y_true, y_pred_prob, title, filename, output_dir='./plots'):
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, filename)

    precision, recall, _ = precision_recall_curve(y_true, y_pred_prob)
    average_precision = average_precision_score(y_true, y_pred_prob)

    plt.figure()
    plt.step(recall, precision, where='post')
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    plt.ylim([0.0, 1.05])
    plt.xlim([0.0, 1.0])
    plt.title(f'{title} - AP: {average_precision:.2f}')
    plt.savefig(filepath)
    plt.close()


# Function to plot feature importance
def plot_feature_importance(importances, feature_names, title, filename, output_dir='./plots'):
    """
    Plot feature importance as a horizontal bar chart, similar to the R code.

    Parameters:
    importances (array-like): Feature importance scores.
    feature_names (array-like): Names of the features.
    title (str): Title of the plot.
    filename (str): Name of the file to save the plot.
    output_dir (str): Directory where the plot will be saved.
    """
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, filename)

    # Create a DataFrame for feature importances
    importance_data = pd.DataFrame({
        'Feature': feature_names,
        'Importance': importances
    })

    # Order the features by absolute importance value for better visualization
    importance_data = importance_data.sort_values(by='Importance', key=abs, ascending=False)

    # Create a horizontal bar plot
    plt.figure(figsize=(10, 8))
    sns.barplot(x='Importance', y='Feature', data=importance_data, palette='Blues_d')
    plt.title(title)
    plt.xlabel('Coefficient')
    plt.ylabel('Feature')
    plt.tight_layout()

    # Save the plot as an image file
    plt.savefig(filepath, dpi=300)
    plt.close()



def plot_shap_values(model, X, feature_names, filename, output_dir='./plots'):
    """
    Plot Shapley values for a given model and dataset.

    Parameters:
    model (object): The trained model (RF, SVM, LR, NB).
    X (DataFrame): The dataset (features) to compute Shapley values.
    feature_names (array-like): List of feature names.
    filename (str): Name of the file to save the plot.
    output_dir (str): Directory where the plot will be saved.
    """
    ensure_directory_exists(output_dir)
    filepath = os.path.join(output_dir, filename)

    feature_names = [f.split('_')[0] for f in feature_names]


    try:
        # Select the appropriate SHAP explainer based on the model type
        if isinstance(model, RandomForestClassifier):
            explainer = shap.TreeExplainer(model)
            shap_values = explainer(X, check_additivity=False)  # Explanation object
            shap_values.feature_names = feature_names
        elif isinstance(model, LogisticRegression):
            explainer = shap.LinearExplainer(model, X)
            shap_values = explainer(X)
        elif isinstance(model, (SVC, GaussianNB)):
            explainer = shap.KernelExplainer(model.predict, X)
            shap_values = explainer(X)
        else:
            raise ValueError("Unsupported model type for SHAP value computation.")

        # Handle SHAP values for binary/multi-class classification
        if isinstance(shap_values.values, list) or len(shap_values.values.shape) > 2:
            shap_values_to_plot = shap_values[..., 1]  # Select SHAP values for the positive class (1)
        else:
            shap_values_to_plot = shap_values

        # Ensure the SHAP values are one-dimensional
        if len(shap_values_to_plot.shape) > 2:
            raise ValueError("The SHAP values passed for plotting should be one-dimensional.")

        plt.figure(figsize=(10, 8))
        shap.plots.beeswarm(shap_values_to_plot, show=False)
        plt.tight_layout()
        plt.savefig(filepath)
        plt.close()

    except Exception as e:
        print(f"Error while plotting SHAP for {filename}: {e}")





def plot_learning_curve(estimator, X, y, title, filename, cv=None, n_jobs=-1, train_sizes=np.linspace(.1, 1.0, 5)):
    """
    Generate a simple plot of the test and training learning curve.

    Parameters:
    estimator (object): An estimator object implementing `fit` and `predict`.
    X (array-like): Training vector.
    y (array-like): Target vector relative to X.
    title (str): Title of the plot.
    filename (str): Filename to save the plot.
    cv (int, cross-validation generator or an iterable): Determines the cross-validation splitting strategy.
    n_jobs (int): Number of jobs to run in parallel (default: -1).
    train_sizes (array-like): Relative or absolute numbers of training examples to be used to generate the learning curve.
    """
    plt.figure()
    plt.title(title)
    plt.xlabel("Training examples")
    plt.ylabel("Score")

    train_sizes, train_scores, test_scores = learning_curve(
        estimator, X, y, cv=cv, n_jobs=n_jobs, train_sizes=train_sizes)

    train_scores_mean = np.mean(train_scores, axis=1)
    train_scores_std = np.std(train_scores, axis=1)
    test_scores_mean = np.mean(test_scores, axis=1)
    test_scores_std = np.std(test_scores, axis=1)

    plt.grid()
    plt.fill_between(train_sizes, train_scores_mean - train_scores_std,
                     train_scores_mean + train_scores_std, alpha=0.1,
                     color="r")
    plt.fill_between(train_sizes, test_scores_mean - test_scores_std,
                     test_scores_mean + test_scores_std, alpha=0.1, color="g")
    plt.plot(train_sizes, train_scores_mean, 'o-', color="r",
             label="Training score")
    plt.plot(train_sizes, test_scores_mean, 'o-', color="g",
             label="Cross-validation score")

    plt.legend(loc="best")
    plt.savefig(filename)
    plt.close()



def save_excel_sheet(df, filepath, sheetname, index=False):
    # Create file if it does not exist
    if not os.path.exists(filepath):
        df.to_excel(filepath, sheet_name=sheetname, index=index)

    # Otherwise, add a sheet. Overwrite if there exists one with the same name.
    else:
        with pd.ExcelWriter(filepath, engine='openpyxl', if_sheet_exists='replace', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheetname, index=index)



def save_classification_results(results, output_file, num_features, method='train_test_split'):
    """
    Save evaluation results to an Excel file.

    Parameters:
    results (dict): Evaluation results for each classifier.
    output_file (str): Path to save the Excel file.
    num_features (int): Number of features used in the classification.
    method (str): Method used for evaluation ('train_test_split' or 'cross_validation').
    """
    print(f"Saving evaluation results to {output_file} using method '{method}' with {num_features} features.")

    if method == 'train_test_split':
        rows = []
        for dataset, classification_results in results.items():
            for classifier, data in classification_results.items():
                metrics = data.get('metrics', {})
                ci = data.get('confidence_intervals', {})
                auc_cv = 0
                #if dataset == 'train': auc_cv = data.get('auc_cv')
                row = [
                    dataset.capitalize(),
                    classifier,
                    f"{metrics.get('roc_auc', 'N/A'):.2f} ({ci.get('roc_auc', ['N/A', 'N/A'])[0]:.2f}, {ci.get('roc_auc', ['N/A', 'N/A'])[1]:.2f})",
                    #f"{auc_cv:.2f}",
                    f"{metrics.get('sensitivity', 'N/A'):.2f} ({ci.get('sensitivity', ['N/A', 'N/A'])[0]:.2f}, {ci.get('sensitivity', ['N/A', 'N/A'])[1]:.2f})",
                    f"{metrics.get('specificity', 'N/A'):.2f} ({ci.get('specificity', ['N/A', 'N/A'])[0]:.2f}, {ci.get('specificity', ['N/A', 'N/A'])[1]:.2f})",
                    f"{metrics.get('ppv', 'N/A'):.2f} ({ci.get('ppv', ['N/A', 'N/A'])[0]:.2f}, {ci.get('ppv', ['N/A', 'N/A'])[1]:.2f})",
                    f"{metrics.get('npv', 'N/A'):.2f} ({ci.get('npv', ['N/A', 'N/A'])[0]:.2f}, {ci.get('npv', ['N/A', 'N/A'])[1]:.2f})",
                ]
                rows.append(row)

        # df = pd.DataFrame(rows, columns=['Dataset', 'Classifier', 'AUC (95% CI)', 'AUC_CV', 'Sensitivity (95% CI)',
        #                                  'Specificity (95% CI)',
        #                                  'PPV (95% CI)', 'NPV (95% CI)'])

        df = pd.DataFrame(rows, columns=['Dataset', 'Classifier', 'AUC (95% CI)', 'Sensitivity (95% CI)',
                                         'Specificity (95% CI)',
                                         'PPV (95% CI)', 'NPV (95% CI)'])

    elif method == 'cross_validation' or method == 'cv_feature_selection_model_building':
        rows = []
        for classifier, data in results.items():
            metrics = data.get('metrics', {})
            ci = data.get('confidence_intervals', {})
            row = [
                classifier,
                f"{metrics.get('roc_auc', 'N/A'):.2f} ({ci.get('roc_auc', ['N/A', 'N/A'])[0]:.2f}, {ci.get('roc_auc', ['N/A', 'N/A'])[1]:.2f})",
                f"{metrics.get('sensitivity', 'N/A'):.2f} ({ci.get('sensitivity', ['N/A', 'N/A'])[0]:.2f}, {ci.get('sensitivity', ['N/A', 'N/A'])[1]:.2f})",
                f"{metrics.get('specificity', 'N/A'):.2f} ({ci.get('specificity', ['N/A', 'N/A'])[0]:.2f}, {ci.get('specificity', ['N/A', 'N/A'])[1]:.2f})",
                f"{metrics.get('ppv', 'N/A'):.2f} ({ci.get('ppv', ['N/A', 'N/A'])[0]:.2f}, {ci.get('ppv', ['N/A', 'N/A'])[1]:.2f})",
                f"{metrics.get('npv', 'N/A'):.2f} ({ci.get('npv', ['N/A', 'N/A'])[0]:.2f}, {ci.get('npv', ['N/A', 'N/A'])[1]:.2f})",
            ]
            rows.append(row)

        df = pd.DataFrame(rows, columns=['Classifier', 'AUC (95% CI)', 'Sensitivity (95% CI)', 'Specificity (95% CI)',
                                         'PPV (95% CI)', 'NPV (95% CI)'])


    else:
        raise ValueError("Invalid method. Choose 'train_test_split' or 'cross_validation'.")

    # Save results
    sheetname = str(num_features) + "_features"
    save_excel_sheet(df, output_file, sheetname)










def cv_feature_selection_model_building_evaluation(x, y, cv_folds=5, tuning=False, result_path="./results",
                                num_features=10, resampling_method=None) -> dict:
    """
    Perform feature selection using MRMR at each fold of cross-validation,
    then build and evaluate models.

    :param df: DataFrame containing features and the outcome variable.
    :param outcome_column: The name of the outcome column.
    :param exclude_columns: List of columns to exclude from the analysis.
    :param num_features: Number of features to select using MRMR.
    :param cv_folds: Number of cross-validation folds.
    :param tuning: Whether to perform hyperparameter tuning.
    :param result_path: Path to save the results.
    :param resampling_method: Resampling method to handle class imbalance.
    :return: Dictionary with evaluation results.
    """


    classifiers = get_classifiers()
    results = {}

    roc_data_path = os.path.join(result_path, "ROC_data")
    ensure_directory_exists(roc_data_path)
    prob_data_path = os.path.join(result_path, "Prob_data")
    ensure_directory_exists(prob_data_path)
    roc_path = os.path.join(result_path, "ROC_curves")
    ensure_directory_exists(roc_path)
    calibration_path = os.path.join(result_path, "Calibration_plots")
    dca_path = os.path.join(result_path, "DCA_curves")
    shap_path = os.path.join(result_path, "Shapley_plots")
    importance_path = os.path.join(result_path, "Feature_Importance_plots")
    model_path = os.path.join(result_path, "Saved_Models")
    ensure_directory_exists(model_path)

    plt.figure(figsize=(10, 8))
    plt.title('ROC Curves', fontsize=20)
    plt.plot([0, 1], [0, 1], color='navy', linestyle='--', lw=2)

    skf = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=17)

    for name, (clf, param_grid) in classifiers.items():
        metrics_list = []
        fpr_list = []
        tpr_list = []
        thresholds_list = []
        auc_list = []
        y_pred_prob_all_folds = []

        if tuning:
            print(f"Hyperparameter_tuning for {name} classifier")
            clf = hyperparameter_tuning(clf, param_grid, x, y, name)

        selected_feature_count = defaultdict(int)
        for feature in x.columns:
            selected_feature_count[feature] = 0

        for fold, (train_index, test_index) in enumerate(skf.split(x, y), 1):
            X_train, X_test = x.iloc[train_index], x.iloc[test_index]
            y_train, y_test = y.iloc[train_index], y.iloc[test_index]

            # Feature selection using MRMR
            selected_features = mrmr_classif(X=X_train, y=y_train, K=num_features)
            X_train = X_train[selected_features]
            X_test = X_test[selected_features]

            for feature in selected_features:
                selected_feature_count[feature] += 1

            if resampling_method:
                X_train, y_train = resampling_method.fit_resample(X_train, y_train)

            clf.fit(X_train, y_train)

            y_pred = clf.predict(X_test)
            y_pred_prob = clf.predict_proba(X_test)[:, 1] if hasattr(clf, "predict_proba") else None
            y_pred_prob_all_folds.append(y_pred_prob)
            metrics, _ = compute_metrics(y_test, y_pred, y_pred_prob, 0.8)
            metrics_list.append(metrics)

            # Collect data for ROC plotting
            fpr, tpr, thresholds = roc_curve(y_test, y_pred_prob)
            roc_auc = roc_auc_score(y_test, y_pred_prob)
            fpr_list.append(fpr)
            tpr_list.append(tpr)
            thresholds_list.append(thresholds)
            auc_list.append(roc_auc)

            # # Plot feature importance for tree-based models
            # if hasattr(clf, 'feature_importances_'):
            #     plot_feature_importance(clf.feature_importances_, X_train.columns, 'Feature Importance',
            #                             f'{name}_featureImportance_{num_features}_features.png',
            #                             output_dir=importance_path)
            #
            # # Plot Shapley values
            # if num_features > 1 and name in ['RandomForest', 'XGBoost', 'LightGBM']:
            #     plot_shap_values(clf, X_test, X_train.columns, f'{name}_shap_values_{num_features}_features.png',
            #                      output_dir=shap_path)



        # Average metrics and confidence intervals across folds
        averaged_metrics = {metric: np.mean([m[metric] for m in metrics_list if m[metric] is not None]) for metric in
                            metrics_list[0]}
        ci = {metric: compute_confidence_interval(averaged_metrics[metric], y.size) for metric in averaged_metrics}

        results[name] = {
            'metrics': averaged_metrics,
            'confidence_intervals': ci
        }

        # Plot and save averaged ROC curve across folds
        mean_fpr = np.linspace(0, 1, 100)
        mean_tpr = np.mean([np.interp(mean_fpr, fpr, tpr) for fpr, tpr in zip(fpr_list, tpr_list)], axis=0)
        mean_tpr[-1] = 1.0
        mean_auc = np.mean(auc_list)

        plt.plot(mean_fpr, mean_tpr, lw=2, label=f'{name} (AUC = {mean_auc:.2f})')

        plot_roc_curve(mean_fpr, mean_tpr, mean_auc, f'{name} Averaged ROC Curve',
                       filename=f'{name}_averaged_roc_curve_{num_features}_features.png',
                       output_dir=os.path.join(roc_path, name))

        # Align the predicted probabilities across folds by stacking them vertically
        all_probs = np.concatenate(y_pred_prob_all_folds)
        probs_df = pd.DataFrame({'y_pred_prob': all_probs})
        new_excel_path = os.path.join(prob_data_path, f'{name}_predicted_probs_{num_features}_features.xlsx')
        probs_df.to_excel(new_excel_path, index=False)

        # Save the averaged ROC data
        roc_data_avg_df = pd.DataFrame({
            'mean_fpr': mean_fpr,
            'mean_tpr': mean_tpr,
        })
        roc_data_avg_df.to_excel(
            os.path.join(roc_data_path, f'{name}_averaged_roc_data_{num_features}_features.xlsx'), index=False)

        # Train the final model on the entire dataset and save it
        selected_feature_count_sorted = {k: v for k, v in sorted(selected_feature_count.items(), key=lambda item: item[1], reverse=True)}
        top_features = list(selected_feature_count_sorted.keys())[:num_features]
        final_model = clf.fit(x[top_features], y)
        final_model.feature_names = top_features
        fnames = final_model.feature_names
        final_model_filename = os.path.join(model_path, f'{name}_{num_features}_features.pkl')
        joblib.dump(final_model, final_model_filename)

    plt.xlabel('False Positive Rate', fontsize=18)
    plt.ylabel('True Positive Rate', fontsize=18)
    plt.legend(loc='lower right', fontsize=16)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    filepath = os.path.join(roc_path, f'averaged_roc_curve_{num_features}_features.png')
    #plt.savefig(filepath, dpi=300)

    return results
